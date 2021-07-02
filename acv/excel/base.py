from openpyxl.styles import Side, Border, colors
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.drawing.image import Image


class Excel:
    def __init__(self):
        self.excel = None
        self.sheet = None

    def open(self, excel_path, sheet_name):
        """
        open an excel file.
        :param excel_path: excel file path
        :param sheet_name: a sheet name
        :return:
        """
        self.excel = openpyxl.load_workbook(excel_path)
        self.sheet = self.excel.get_sheet_by_name(sheet_name)

    def get_value(self, col: str, row: int):
        """
        get a value from sheet
        :param col: col coord
        :param row: row coord
        :return: value
        """
        return self.sheet[col][row]

    def save(self, save_path):
        """
        save excel file.
        :param save_path: save path
        :return:
        """
        self.excel.save(save_path)

    def add_image(self, img_path, col: str, row: int):
        """
        add image to the sheet
        :param img_path: image file path
        :param col: col coord
        :param row: row coord
        :return:
        """
        ws = self.excel.active
        img = Image(img_path)
        img.width, img.height = img.width / 2, img.height / 2
        ws.add_image(img, col + str(row))

    def set_color(self, col, row, color):
        """
        :param col: col number, e.g: A,B,C.....
        :param row: row number, e.g: 119
        :param color: 'BDD7EE', 'F4B084'
        :return:
        """
        self.sheet[col][row].fill = PatternFill('solid', fgColor=color)

    def set_value(self, col: str, row: int, value):
        """
        modify a value of sheet
        :param col: col coord
        :param row: row coord
        :param value: input value
        :return:
        """
        self.sheet[col][row].value = value

    @staticmethod
    def border(t_border='thin', b_border='thin', l_border='thin', r_border='thin'):
        """
        :param t_border: side style('dashDot','dashDotDot', 'dashed','dotted',
                            'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                            'mediumDashed', 'slantDashDot', 'thick', 'thin')
        :param b_border: side style
        :param l_border: side style
        :param r_border: side style
        :return:
        """
        border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                        bottom=Side(border_style=b_border, color=colors.BLACK),
                        left=Side(border_style=l_border, color=colors.BLACK),
                        right=Side(border_style=r_border, color=colors.BLACK))
        return border

    def set_border(self, col: str, row: int, style=['thin', 'thin', 'thin', 'thin']):
        """
        set border style
        :param col: col coord
        :param row: row coord
        :param style: border style
        :return:
        """
        self.sheet[col][row].border = self.border(style[0], style[0], style[0], style[0])


if __name__ == '__main__':
    e = Excel()
    e.open("./empty.xlsx", 'Sheet2')
    e.set_value('A', 2, '1')
    e.set_value('A', 3, '778')
    e.save('./new_.xlsx')
